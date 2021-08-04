import requests
import openpyxl
from openpyxl import cell
import whois
from dns import resolver

domainfile=openpyxl.load_workbook("domain-report.xlsx")

domains=domainfile["Sheet1"]

for i in range(2, domains.max_row + 1):
    domain=domains.cell(i,1).value
    print(domain)
    
    Arecord = resolver.resolve(domain)
    record=""
    for ipval in Arecord:
        print(ipval.to_text())
        record=record+str(ipval)+"\n"
        
    domains.cell(i,5).value=record
    MXrecord = resolver.resolve(domain,'mx')
    record=""
    for ipval in MXrecord:
        print(ipval.to_text())
        record=record+str(ipval)+"\n"
    domains.cell(i,6).value=record
    
    TXTrecord = resolver.resolve(domain,'txt')
    record=""
    for ipval in TXTrecord:
        print(ipval.to_text())
        record=record+str(ipval)+"\n"
    domains.cell(i,7).value=record
    
    dwhois=whois.whois(domain)
    #print(dwhois)
    #domains.cell(i,4).value=
   # print(dwhois['expiration_date'][1])
    #print(dwhois['expiration_date'][1].strftime('%m/%d/%Y'))
    domains.cell(i,3).value=str(dwhois['registrar'])
    try:
        expdate=str(dwhois['expiration_date'][1])
        domains.cell(i,4).value=expdate
    except:
        print("no")
    
    rcode=str(requests.head('http://'+domain+'', headers={'User-Agent': 'Foo bar'},allow_redirects=True).status_code)
    domains.cell(i,2).value=rcode
    
    try:
        domains.cell(i,8).value = str(resolver.resolve("default._domainkey."+domain,'txt'))
    except:
        print("no")
    
    

    
domainfile.save("solution.xlsx")    